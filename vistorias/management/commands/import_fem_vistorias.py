from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from ptms.models import PTM
from vistorias.models import VistoriaPTM


def _norm(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).strip().lower()


def _sheet_by_name(workbook, expected: str) -> Worksheet:
    expected_norm = _norm(expected)
    for ws in workbook.worksheets:
        if _norm(ws.title) == expected_norm:
            return ws
    raise CommandError(f"Aba '{expected}' nao encontrada no arquivo.")


def _to_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted.date()
            if isinstance(converted, date):
                return converted
        except (OverflowError, ValueError, TypeError):
            return None
    return None


def _to_percentage(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0.0000")
    try:
        return Decimal(str(value)).quantize(Decimal("0.0001"))
    except (InvalidOperation, ValueError):
        return Decimal("0.0000")


def _iter_rows(ws: Worksheet):
    for row in range(2, ws.max_row + 1):
        ordem = _to_str(ws[f"A{row}"].value)
        if ordem:
            yield row, ordem


@dataclass
class Counters:
    processed: int = 0
    vistorias: int = 0
    ptm_missing: int = 0


class Command(BaseCommand):
    help = "Importa apenas vistorias da planilha FEM para PTMs ja existentes."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Caminho da planilha .xlsx/.xlsm")
        parser.add_argument("--limit", type=int, default=0, help="Limite de linhas para teste")

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"Arquivo nao encontrado: {file_path}")

        limit = max(0, int(options["limit"] or 0))
        wb = load_workbook(file_path, data_only=True, keep_vba=True)
        ws = _sheet_by_name(wb, "VISTORIA")

        counters = Counters()
        for row, ordem in _iter_rows(ws):
            if limit and counters.processed >= limit:
                break

            ptm = PTM.objects.filter(ordem=ordem).first()
            if ptm is None:
                counters.ptm_missing += 1
                counters.processed += 1
                continue

            with transaction.atomic():
                ptm.vistorias.all().delete()
                counters.vistorias += self._import_vistorias(ws, row, ptm)

            counters.processed += 1
            if counters.processed % 25 == 0:
                self.stdout.write(f"Processados {counters.processed} PTMs...")

        self.stdout.write(self.style.SUCCESS("Importacao de vistorias concluida."))
        self.stdout.write(
            f"Linhas processadas: {counters.processed} | vistorias criadas: {counters.vistorias}"
        )
        if counters.ptm_missing:
            self.stdout.write(f"PTMs nao encontrados (ignorados): {counters.ptm_missing}")

    def _import_vistorias(self, ws: Worksheet, row: int, ptm: PTM) -> int:
        created = 0
        col = 6  # F
        idx = 1
        while col + 3 <= ws.max_column:
            c_sol = get_column_letter(col)
            c_resp = get_column_letter(col + 1)
            c_pct = get_column_letter(col + 2)
            c_obs = get_column_letter(col + 3)
            if ws[f"{c_sol}{row}"].value in (None, "") and ws[f"{c_resp}{row}"].value in (None, ""):
                col += 4
                idx += 1
                continue
            VistoriaPTM.objects.create(
                ptm=ptm,
                ordem_vistoria=idx,
                dt_solicitacao=_to_date(ws[f"{c_sol}{row}"].value),
                dt_resposta=_to_date(ws[f"{c_resp}{row}"].value),
                percentual_execucao=_to_percentage(ws[f"{c_pct}{row}"].value),
                observacao=_to_str(ws[f"{c_obs}{row}"].value),
            )
            created += 1
            col += 4
            idx += 1
        return created
