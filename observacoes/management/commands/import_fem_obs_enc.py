from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet

from observacoes.models import ObservacaoEncaminhamentoPTM
from ptms.models import PTM


def _norm(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).strip().lower()


def _sheet_by_name(workbook, expected: str) -> Worksheet:
    expected_norm = _norm(expected)
    for ws in workbook.worksheets:
        if _norm(ws.title) == expected_norm:
            return ws
    raise CommandError(f"Aba '{expected}' nao encontrada no arquivo.")


def _to_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted.date()
            if isinstance(converted, date):
                return converted
        except (OverflowError, ValueError, TypeError):
            return None
    return None


def _iter_rows(ws: Worksheet):
    for row in range(2, ws.max_row + 1):
        ordem = _to_str(ws[f"A{row}"].value)
        if ordem:
            yield row, ordem


def _extract_origem(header_value: str) -> str:
    header = _to_str(header_value)
    if not header:
        return ""
    match = re.search(r"\((.*?)\)", header)
    if match:
        return match.group(1).strip()
    return ""


@dataclass
class Counters:
    processed: int = 0
    observacoes: int = 0
    ptm_missing: int = 0


class Command(BaseCommand):
    help = "Importa observacoes/encaminhamentos da aba OBS  ENC para PTMs existentes."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Caminho da planilha .xlsx/.xlsm")
        parser.add_argument("--limit", type=int, default=0, help="Limite de linhas para teste")
        parser.add_argument(
            "--max-pares",
            type=int,
            default=0,
            help="Limite de pares OBS/DATA por linha (0 = sem limite).",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"Arquivo nao encontrado: {file_path}")

        limit = max(0, int(options["limit"] or 0))
        max_pairs = max(0, int(options["max_pares"] or 0))
        wb = load_workbook(file_path, data_only=True, keep_vba=True)
        ws = _sheet_by_name(wb, "OBS  ENC")

        counters = Counters()
        for row, ordem in _iter_rows(ws):
            if limit and counters.processed >= limit:
                break

            ptm = PTM.objects.filter(ordem=ordem).first()
            if ptm is None:
                counters.ptm_missing += 1
                counters.processed += 1
                continue

            with transaction.atomic():
                ptm.observacoes_enc.all().delete()
                counters.observacoes += self._import_obs(ws, row, ptm, max_pairs=max_pairs)

            counters.processed += 1
            if counters.processed % 25 == 0:
                self.stdout.write(f"Processados {counters.processed} PTMs...")

        self.stdout.write(self.style.SUCCESS("Importacao de OBS/ENC concluida."))
        self.stdout.write(
            f"Linhas processadas: {counters.processed} | observacoes criadas: {counters.observacoes}"
        )
        if counters.ptm_missing:
            self.stdout.write(f"PTMs nao encontrados (ignorados): {counters.ptm_missing}")

    def _import_obs(self, ws: Worksheet, row: int, ptm: PTM, max_pairs: int = 0) -> int:
        created = 0
        col = 6  # F
        pair_idx = 0
        while col + 1 <= ws.max_column:
            if max_pairs and pair_idx >= max_pairs:
                break
            c_obs = get_column_letter(col)
            c_data = get_column_letter(col + 1)
            obs = _to_str(ws[f"{c_obs}{row}"].value)
            data = _to_date(ws[f"{c_data}{row}"].value)
            if obs or data:
                origem = _extract_origem(ws[f"{c_obs}1"].value)
                ObservacaoEncaminhamentoPTM.objects.create(
                    ptm=ptm,
                    data=data,
                    observacao=obs or "(sem observacao)",
                    origem=origem,
                )
                created += 1
            col += 2
            pair_idx += 1
        return created
