from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet

from prestacao_contas.models import PrestacaoContaHistorico, PrestacaoContaPTM
from ptms.models import PTM


def _norm(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).strip().lower()


def _sheet_by_name(workbook, expected: str) -> Worksheet:
    expected_norm = _norm(expected)
    for ws in workbook.worksheets:
        if _norm(ws.title) == expected_norm:
            return ws
    raise CommandError(f"Aba '{expected}' nao encontrada no arquivo.")


def _to_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted.date()
            if isinstance(converted, date):
                return converted
        except (OverflowError, ValueError, TypeError):
            return None
    return None


def _iter_rows(ws: Worksheet):
    for row in range(2, ws.max_row + 1):
        ordem = _to_str(ws[f"A{row}"].value)
        if ordem:
            yield row, ordem


@dataclass
class Counters:
    processed: int = 0
    prestacoes: int = 0
    historicos: int = 0
    ptm_missing: int = 0


class Command(BaseCommand):
    help = "Importa prestacao de contas da planilha FEM para PTMs ja existentes."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Caminho da planilha .xlsx/.xlsm")
        parser.add_argument("--limit", type=int, default=0, help="Limite de linhas para teste")
        parser.add_argument(
            "--max-historico-pares",
            type=int,
            default=0,
            help="Limite de pares OBS/DATA por linha (0 = sem limite).",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"Arquivo nao encontrado: {file_path}")

        limit = max(0, int(options["limit"] or 0))
        max_pairs = max(0, int(options["max_historico_pares"] or 0))
        wb = load_workbook(file_path, data_only=True, keep_vba=True)
        ws = _sheet_by_name(wb, "PRESTACAO DE CONTAS")

        counters = Counters()
        for row, ordem in _iter_rows(ws):
            if limit and counters.processed >= limit:
                break

            ptm = PTM.objects.filter(ordem=ordem).first()
            if ptm is None:
                counters.ptm_missing += 1
                counters.processed += 1
                continue

            with transaction.atomic():
                PrestacaoContaPTM.objects.filter(ptm=ptm).delete()
                prest_count, hist_count = self._import_prestacao(ws, row, ptm, max_pairs=max_pairs)
                counters.prestacoes += prest_count
                counters.historicos += hist_count

            counters.processed += 1
            if counters.processed % 25 == 0:
                self.stdout.write(f"Processados {counters.processed} PTMs...")

        self.stdout.write(self.style.SUCCESS("Importacao de prestacao de contas concluida."))
        self.stdout.write(
            f"Linhas processadas: {counters.processed} | prestacoes criadas: {counters.prestacoes} | historicos criados: {counters.historicos}"
        )
        if counters.ptm_missing:
            self.stdout.write(f"PTMs nao encontrados (ignorados): {counters.ptm_missing}")

    def _import_prestacao(self, ws: Worksheet, row: int, ptm: PTM, max_pairs: int = 0) -> tuple[int, int]:
        prazo_contas = _to_date(ws[f"F{row}"].value)
        data_prestacao = _to_date(ws[f"G{row}"].value)
        situacao = _to_str(ws[f"H{row}"].value)

        has_hist = False
        col = 9  # I
        pairs_scanned = 0
        while col + 1 <= ws.max_column:
            if max_pairs and pairs_scanned >= max_pairs:
                break
            obs = _to_str(ws.cell(row=row, column=col).value)
            data_item = _to_date(ws.cell(row=row, column=col + 1).value)
            if obs or data_item:
                has_hist = True
                break
            col += 2
            pairs_scanned += 1

        if not (prazo_contas or data_prestacao or situacao or has_hist):
            return 0, 0

        prestacao = PrestacaoContaPTM.objects.create(
            ptm=ptm,
            prazo_contas=prazo_contas,
            data_prestacao=data_prestacao,
            situacao=situacao,
        )

        hist_count = 0
        col = 9  # I
        pair_idx = 0
        while col + 1 <= ws.max_column:
            if max_pairs and pair_idx >= max_pairs:
                break
            c_obs = get_column_letter(col)
            c_data = get_column_letter(col + 1)
            obs = _to_str(ws[f"{c_obs}{row}"].value)
            data_item = _to_date(ws[f"{c_data}{row}"].value)
            if obs or data_item:
                PrestacaoContaHistorico.objects.create(
                    prestacao=prestacao,
                    data=data_item,
                    observacao=obs or "(sem observacao)",
                )
                hist_count += 1
            col += 2
            pair_idx += 1

        return 1, hist_count
