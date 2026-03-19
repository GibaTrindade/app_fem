from decimal import Decimal, InvalidOperation
from functools import lru_cache
import json
from pathlib import Path
import unicodedata
from urllib.parse import urlencode

from django.contrib import messages
from django.db import transaction
from django.db.models import Count, DateField, DecimalField, F, Max, OuterRef, Q, Subquery, Sum, Value
from django.db.models.functions import Coalesce, Substr
from django.http import Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.generic import CreateView, DeleteView, DetailView, ListView, TemplateView, UpdateView
from openpyxl import load_workbook

from conclusao_informal.models import ConclusaoInformalPTM
from core.models import StatusObra, StatusPTM, TermoAdesaoObservacao, TermoAdesaoResponsavel
from eventos.models import EventoPTM
from observacoes.models import ObservacaoEncaminhamentoPTM
from pagamentos.models import PagamentoPTM
from prestacao_contas.models import PrestacaoContaHistorico, PrestacaoContaPTM
from ptms.forms import (
    ConclusaoInformalForm,
    DocumentoPublicoPTMForm,
    EventoPTMForm,
    NotaTecnicaMunicipioForm,
    ObservacaoEncaminhamentoForm,
    PagamentoPTMForm,
    PrestacaoContaHistoricoForm,
    PrestacaoContaPTMForm,
    PTMForm,
    StatusAnaliseDocumentacaoForm,
    TermoAdesaoPTMForm,
    VistoriaPTMForm,
)
from ptms.models import NotaTecnicaPTM, PTM, TermoAdesaoPTM
from vistorias.models import VistoriaPTM


def _normalize_text(value):
    text = (value or "").strip()
    text = "".join(
        char for char in unicodedata.normalize("NFD", text) if unicodedata.category(char) != "Mn"
    )
    return " ".join(text.lower().split())


def _normalize_municipio(value):
    return _normalize_text(value)


def _allowed_municipios_for_user(user):
    if user.is_superuser:
        return None
    return {
        _normalize_municipio(v)
        for v in user.user_municipios.select_related("municipio").values_list("municipio__nome", flat=True)
    }


def _user_can_edit_ptm(user, ptm):
    allowed = _allowed_municipios_for_user(user)
    if allowed is None:
        return True
    if not allowed:
        return False
    return _normalize_municipio(ptm.municipio) in allowed


def _deny_edit_if_forbidden(request, ptm, tab="eventos"):
    if _user_can_edit_ptm(request.user, ptm):
        return None
    messages.error(
        request,
        "Voce nao tem permissao para alterar este PTM. Verifique seus municipios vinculados.",
    )
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab={tab}")


def _public_ptm_url(request, ptm):
    if not ptm.codigo_acesso_publico:
        return ""
    return request.build_absolute_uri(
        reverse("ptm_public_detail", kwargs={"token": ptm.codigo_acesso_publico})
    )


def _municipios_com_ptms_para_usuario(user):
    municipios = sorted(set(PTM.objects.order_by("municipio").values_list("municipio", flat=True)))
    allowed = _allowed_municipios_for_user(user)
    if allowed is None:
        return municipios
    return [municipio for municipio in municipios if _normalize_municipio(municipio) in allowed]


def _resolve_municipio_para_nota(user, municipio_raw):
    municipio_norm = _normalize_municipio(municipio_raw)
    if not municipio_norm:
        return ""
    for municipio in _municipios_com_ptms_para_usuario(user):
        if _normalize_municipio(municipio) == municipio_norm:
            return municipio
    return ""


def _nota_tecnica_default_observacao(ptm):
    ultimo_evento = ptm.eventos.order_by("-data_evento", "-id").first()
    if ultimo_evento and ultimo_evento.descricao:
        return ultimo_evento.descricao

    prestacao = getattr(ptm, "prestacao_conta", None)
    if prestacao:
        historico = prestacao.historico.order_by("-data", "-id").first()
        if historico and historico.observacao:
            return historico.observacao
        if prestacao.situacao:
            return prestacao.situacao

    ultima_obs = ptm.observacoes_enc.order_by("-data", "-id").first()
    if ultima_obs and ultima_obs.observacao:
        return ultima_obs.observacao

    if ptm.ressalva:
        return ptm.ressalva
    if ptm.status_ptm_atual:
        return ptm.status_ptm_atual.nome
    return ""


def _nota_tecnica_observacao_atual(ptm):
    try:
        nota = ptm.nota_tecnica
    except NotaTecnicaPTM.DoesNotExist:
        nota = None
    if nota is not None:
        return nota.observacao
    return _nota_tecnica_default_observacao(ptm)


def _nota_tecnica_secao_label(ptm):
    tipo_nome = (ptm.tipo_fem.nome or "").upper() if ptm.tipo_fem_id else ""
    if tipo_nome == "EMENDA":
        return "emenda_parlamentar"
    if ptm.ordem.startswith("2013."):
        return "fem_2013"
    if ptm.ordem.startswith("2014."):
        return "fem_2014"
    if ptm.ordem.startswith("2015."):
        return "fem_2015"
    return "outros"


def _nota_tecnica_secao_defs():
    return [
        {"key": "fem_2013", "titulo": "1. FEM 2013", "tipo": "fem", "ano": "2013", "coluna_valor": "VALOR TOTAL FEM"},
        {"key": "fem_2014", "titulo": "2. FEM 2014", "tipo": "fem", "ano": "2014", "coluna_valor": "VALOR TOTAL FEM"},
        {"key": "fem_2015", "titulo": "3. FEM 2015", "tipo": "fem", "ano": "2015", "coluna_valor": "VALOR TOTAL FEM"},
        {
            "key": "emenda_parlamentar",
            "titulo": "4. Emenda Parlamentar",
            "tipo": "emenda",
            "ano": "",
            "coluna_valor": "VALOR UTILIZADO DA EMENDA",
        },
    ]


def _nota_tecnica_valor_total(ptm):
    if ptm.recurso_fem and ptm.recurso_fem > 0:
        return ptm.recurso_fem
    if ptm.investimento_total and ptm.investimento_total > 0:
        return ptm.investimento_total
    return ptm.teto_fem


def _nota_tecnica_texto_abertura(section, total_itens, total_fem, total_repassado):
    total_fem_fmt = f"R$ {total_fem:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    total_repassado_fmt = f"R$ {total_repassado:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    plano_label = "Plano de Trabalho" if total_itens == 1 else "Planos de Trabalho"
    if section["tipo"] == "emenda":
        return (
            f"Para Emendas Parlamentares foram protocolados {total_itens:02d} ({total_itens}) {plano_label}, "
            f"os quais correspondem a um investimento total de {total_fem_fmt}, com {total_repassado_fmt} ja repassado, "
            "abaixo relacionados:"
        )
    if total_itens == 0:
        return f"Ainda nao ha PTMs importados para o bloco {section['titulo']} deste municipio."
    if total_repassado >= total_fem and total_fem > 0:
        repasse_txt = f"{total_fem_fmt} integralmente pago"
    else:
        repasse_txt = f"{total_repassado_fmt} ja repassado"
    return (
        f"Para o FEM de {section['ano']} foram protocolados {total_itens:02d} ({total_itens}) {plano_label}, "
        f"os quais correspondem a um investimento total do FEM de {total_fem_fmt}, com {repasse_txt}, abaixo relacionados:"
    )


def _build_nota_tecnica_context(municipio):
    ptms = (
        PTM.objects.filter(municipio=municipio)
        .select_related("tipo_fem", "status_ptm_atual", "status_obra_atual")
        .prefetch_related(
            "pagamentos",
            "eventos",
            "observacoes_enc",
            "prestacao_conta__historico",
            "nota_tecnica",
        )
        .order_by("ordem")
    )
    grouped = {section["key"]: [] for section in _nota_tecnica_secao_defs()}
    for ptm in ptms:
        section_key = _nota_tecnica_secao_label(ptm)
        if section_key not in grouped:
            continue
        pagamentos = list(ptm.pagamentos.all())
        total_repassado = sum((pag.valor_realizado for pag in pagamentos), Decimal("0.00"))
        ultimo_pagamento = None
        datas_pagamento = [pag.dt_pagamento for pag in pagamentos if pag.dt_pagamento]
        if datas_pagamento:
            ultimo_pagamento = max(datas_pagamento)
        grouped[section_key].append(
            {
                "ptm": ptm,
                "valor_total": _nota_tecnica_valor_total(ptm),
                "valor_repassado": total_repassado,
                "data_ultimo_pagamento": ultimo_pagamento,
                "status": (
                    ptm.status_obra_atual.nome
                    if ptm.status_obra_atual_id
                    else (ptm.status_ptm_atual.nome if ptm.status_ptm_atual_id else "")
                ),
                "observacao": _nota_tecnica_observacao_atual(ptm),
            }
        )

    sections = []
    for section in _nota_tecnica_secao_defs():
        items = grouped[section["key"]]
        total_fem = sum((item["valor_total"] for item in items), Decimal("0.00"))
        total_repassado = sum((item["valor_repassado"] for item in items), Decimal("0.00"))
        sections.append(
            {
                **section,
                "items": items,
                "total_itens": len(items),
                "total_fem": total_fem,
                "total_repassado": total_repassado,
                "texto_abertura": _nota_tecnica_texto_abertura(
                    section,
                    len(items),
                    total_fem,
                    total_repassado,
                ),
            }
        )
    return {"municipio": municipio, "sections": sections}


def _build_dashboard_context():
    today = timezone.localdate()
    stale_limit = today - timezone.timedelta(days=90)

    latest_vistoria = VistoriaPTM.objects.filter(ptm=OuterRef("pk")).order_by("-ordem_vistoria", "-id")
    latest_evento = EventoPTM.objects.filter(ptm=OuterRef("pk")).order_by("-data_evento", "-id")

    kpi_base = PTM.objects.annotate(
        total_pago=Coalesce(
            Sum("pagamentos__valor_realizado"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=14, decimal_places=2),
        ),
        parcelas_count=Count("pagamentos__parcela", distinct=True),
        ultima_vistoria_resposta=Subquery(
            latest_vistoria.values("dt_resposta")[:1],
            output_field=DateField(),
        ),
        ultima_vistoria_solicitacao=Subquery(
            latest_vistoria.values("dt_solicitacao")[:1],
            output_field=DateField(),
        ),
        ultimo_evento=Subquery(
            latest_evento.values("data_evento")[:1],
            output_field=DateField(),
        ),
    )

    total_ptms = PTM.objects.count()
    total_teto_fem = sum(
        (
            item["teto_base"] or Decimal("0.00")
            for item in PTM.objects.exclude(teto_fem__isnull=True)
            .annotate(ano_base=Substr("ordem", 1, 4))
            .values("municipio", "ano_base")
            .annotate(teto_base=Max("teto_fem"))
        ),
        Decimal("0.00"),
    )
    total_pago = kpi_base.aggregate(
        total=Coalesce(
            Sum("total_pago"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=16, decimal_places=2),
        )
    )["total"]
    total_saldo = total_teto_fem - total_pago
    execucao_financeira_pct = (
        ((total_pago / total_teto_fem) * Decimal("100")) if total_teto_fem else Decimal("0.00")
    )
    execucao_financeira_pct = execucao_financeira_pct.quantize(Decimal("0.01"))

    sem_vistoria_qs = PTM.objects.filter(vistorias__isnull=True).order_by("ordem").distinct()

    return {
        "kpi_total_ptms": total_ptms,
        "kpi_total_teto_fem": total_teto_fem,
        "kpi_total_pago": total_pago,
        "kpi_total_saldo": total_saldo,
        "kpi_execucao_financeira_pct": execucao_financeira_pct,
        "kpi_pagamento_pendente": kpi_base.filter(
            Q(parcelas_count=4) & Q(total_pago__lt=F("teto_fem"))
        ).count(),
        "kpi_sem_vistoria": sem_vistoria_qs.count(),
        "sem_vistoria_ptms": sem_vistoria_qs,
        "kpi_vistoria_desatualizada": kpi_base.filter(vistorias__isnull=False)
        .filter(
            Q(ultima_vistoria_resposta__lt=stale_limit)
            | (Q(ultima_vistoria_resposta__isnull=True) & Q(ultima_vistoria_solicitacao__lt=stale_limit))
        )
        .distinct()
        .count(),
        "kpi_prestacao_vencida": PrestacaoContaPTM.objects.filter(prazo_contas__lt=today)
        .filter(Q(data_prestacao__isnull=True) | Q(data_prestacao__gt=F("prazo_contas")))
        .count(),
        "status_breakdown": PTM.objects.values("status_ptm_atual__nome")
        .annotate(total=Count("id"))
        .order_by("-total", "status_ptm_atual__nome")[:6],
        "secretaria_breakdown": PTM.objects.values("secretaria__nome")
        .annotate(total=Count("id"))
        .order_by("-total", "secretaria__nome")[:6],
        "municipio_breakdown": PTM.objects.values("municipio")
        .annotate(total=Count("id"))
        .order_by("-total", "municipio")[:6],
        "alertas_prestacao": PrestacaoContaPTM.objects.select_related("ptm")
        .filter(prazo_contas__lt=today)
        .filter(Q(data_prestacao__isnull=True) | Q(data_prestacao__gt=F("prazo_contas")))
        .order_by("prazo_contas")[:5],
        "alertas_vistoria": kpi_base.filter(vistorias__isnull=False)
        .filter(
            Q(ultima_vistoria_resposta__lt=stale_limit)
            | (Q(ultima_vistoria_resposta__isnull=True) & Q(ultima_vistoria_solicitacao__lt=stale_limit))
        )
        .order_by("ultima_vistoria_resposta", "ultima_vistoria_solicitacao", "ordem")
        .distinct()[:5],
        "alertas_saldo": kpi_base.filter(total_pago__lt=F("teto_fem"))
        .annotate(
            saldo_receber=F("teto_fem") - F("total_pago"),
        )
        .order_by("-saldo_receber", "ordem")[:5],
    }


def _to_decimal_or_none(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _to_date_or_value(value):
    if hasattr(value, "date"):
        try:
            return value.date()
        except TypeError:
            return value
    return value


@lru_cache(maxsize=1)
def _load_resumo_financeiro_from_workbooks():
    cache_path = Path.cwd() / "ptms" / "data" / "resumo_financeiro_cache.json"
    if cache_path.exists():
        try:
            with cache_path.open("r", encoding="utf-8") as fp:
                raw = json.load(fp)
            resumo_por_ordem = {}
            for ordem, item in raw.items():
                resumo_por_ordem[ordem] = {
                    "data_ultimo_pagamento": (
                        timezone.datetime.fromisoformat(item["data_ultimo_pagamento"]).date()
                        if item.get("data_ultimo_pagamento")
                        else None
                    ),
                    "parcela_paga": item.get("parcela_paga") or "SEM PAGAMENTO",
                    "valor_ultimo_pagamento": _to_decimal_or_none(item.get("valor_ultimo_pagamento")),
                    "repasse_total": _to_decimal_or_none(item.get("repasse_total")) or Decimal("0.00"),
                    "repasse_valido": _to_decimal_or_none(item.get("repasse_valido")) or Decimal("0.00"),
                    "percentual_recebido": _to_decimal_or_none(item.get("percentual_recebido")),
                    "saldo_a_receber": _to_decimal_or_none(item.get("saldo_a_receber")) or Decimal("0.00"),
                    "parcela_pendente": item.get("parcela_pendente") or "OK",
                    "situacao": item.get("situacao") or "OK",
                    "data_referencia": (
                        timezone.datetime.fromisoformat(item["data_referencia"]).date()
                        if item.get("data_referencia") not in (None, "", "OK")
                        else (item.get("data_referencia") or "OK")
                    ),
                }
            return resumo_por_ordem
        except (OSError, ValueError, TypeError):
            pass

    resumo_por_ordem = {}
    for path in sorted(Path.cwd().glob("CONTROLE FEM *.xls*")):
        try:
            wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
        except Exception:
            continue
        if "RESUMO" not in wb.sheetnames:
            continue
        ws = wb["RESUMO"]
        for row in range(7, ws.max_row + 1):
            ordem = ws.cell(row=row, column=2).value
            if not ordem:
                continue
            resumo_por_ordem[str(ordem).strip()] = {
                "data_ultimo_pagamento": _to_date_or_value(ws.cell(row=row, column=14).value),
                "parcela_paga": ws.cell(row=row, column=15).value or "SEM PAGAMENTO",
                "valor_ultimo_pagamento": _to_decimal_or_none(ws.cell(row=row, column=16).value),
                "repasse_total": _to_decimal_or_none(ws.cell(row=row, column=17).value) or Decimal("0.00"),
                "repasse_valido": _to_decimal_or_none(ws.cell(row=row, column=18).value) or Decimal("0.00"),
                "percentual_recebido": (
                    _to_decimal_or_none(Decimal(str(ws.cell(row=row, column=19).value)) * Decimal("100"))
                    if ws.cell(row=row, column=19).value not in (None, "", "NA")
                    else None
                ),
                "saldo_a_receber": _to_decimal_or_none(ws.cell(row=row, column=20).value) or Decimal("0.00"),
                "parcela_pendente": ws.cell(row=row, column=21).value or "OK",
                "situacao": ws.cell(row=row, column=22).value or "OK",
                "data_referencia": _to_date_or_value(ws.cell(row=row, column=23).value),
            }
    if resumo_por_ordem:
        try:
            cache_path.parent.mkdir(parents=True, exist_ok=True)
            serializable = {}
            for ordem, item in resumo_por_ordem.items():
                serializable[ordem] = {
                    "data_ultimo_pagamento": (
                        item["data_ultimo_pagamento"].isoformat() if hasattr(item["data_ultimo_pagamento"], "isoformat") else None
                    ),
                    "parcela_paga": item["parcela_paga"],
                    "valor_ultimo_pagamento": (
                        str(item["valor_ultimo_pagamento"]) if item["valor_ultimo_pagamento"] is not None else None
                    ),
                    "repasse_total": str(item["repasse_total"]),
                    "repasse_valido": str(item["repasse_valido"]),
                    "percentual_recebido": (
                        str(item["percentual_recebido"]) if item["percentual_recebido"] is not None else None
                    ),
                    "saldo_a_receber": str(item["saldo_a_receber"]),
                    "parcela_pendente": item["parcela_pendente"],
                    "situacao": item["situacao"],
                    "data_referencia": (
                        item["data_referencia"].isoformat() if hasattr(item["data_referencia"], "isoformat") else item["data_referencia"]
                    ),
                }
            with cache_path.open("w", encoding="utf-8") as fp:
                json.dump(serializable, fp, ensure_ascii=False)
        except OSError:
            pass
    return resumo_por_ordem


def _build_resumo_financeiro(ptm):
    resumo_planilha = _load_resumo_financeiro_from_workbooks().get(ptm.ordem)
    if resumo_planilha:
        return resumo_planilha

    pagamentos = list(ptm.pagamentos.order_by("dt_pagamento", "id"))
    pagamentos_com_data = [pag for pag in pagamentos if pag.dt_pagamento]
    ultimo_pagamento = pagamentos_com_data[-1] if pagamentos_com_data else None

    repasse_total = sum((pag.valor_realizado for pag in pagamentos), Decimal("0.00"))
    repasse_valido = repasse_total
    recurso_fem = ptm.recurso_fem or Decimal("0.00")
    percentual_recebido = None
    if recurso_fem:
        percentual_recebido = ((repasse_valido / recurso_fem) * Decimal("100")).quantize(Decimal("0.01"))
    saldo_a_receber = max(recurso_fem - repasse_valido, Decimal("0.00"))

    pagamentos_normais = {pag.parcela: pag for pag in pagamentos if pag.tipo_registro == "normal"}
    parcela_pendente = "OK"
    situacao = "OK"
    data_referencia = "OK"
    for parcela in ("4", "3", "2", "1"):
        pagamento = pagamentos_normais.get(parcela)
        if pagamento is None:
            continue
        if pagamento.dt_pagamento or not (pagamento.dt_solicitacao or pagamento.dt_envio_pg):
            continue
        parcela_pendente = f"{parcela}a"
        data_referencia = pagamento.dt_envio_pg or pagamento.dt_solicitacao
        if pagamento.dt_envio_pg:
            situacao = "NO FINANCEIRO"
        elif parcela == "1":
            situacao = "AGUARDANDO LIBERACAO"
        else:
            situacao = "AGUARDANDO VISTORIA"
        break

    if ultimo_pagamento:
        if ultimo_pagamento.tipo_registro == "extra":
            parcela_paga = f"{ultimo_pagamento.parcela}a P COMP."
        else:
            parcela_paga = f"{ultimo_pagamento.parcela}a P"
    else:
        parcela_paga = "SEM PAGAMENTO"

    return {
        "data_ultimo_pagamento": ultimo_pagamento.dt_pagamento if ultimo_pagamento else None,
        "parcela_paga": parcela_paga,
        "valor_ultimo_pagamento": ultimo_pagamento.valor_realizado if ultimo_pagamento else None,
        "repasse_total": repasse_total,
        "repasse_valido": repasse_valido,
        "percentual_recebido": percentual_recebido,
        "saldo_a_receber": saldo_a_receber,
        "parcela_pendente": parcela_pendente,
        "situacao": situacao,
        "data_referencia": data_referencia,
    }


class PTMListView(ListView):
    model = PTM
    template_name = "ptms/ptm_list.html"
    context_object_name = "ptms"
    paginate_by = 20

    def get_queryset(self):
        queryset = (
            PTM.objects.select_related("tipo_fem", "status_ptm_atual", "status_obra_atual", "secretaria")
            .order_by("ordem")
        )
        q = self.request.GET.get("q", "").strip()
        status = self.request.GET.get("status", "").strip()
        municipio = self.request.GET.get("municipio", "").strip()

        if q:
            q_norm = _normalize_text(q)
            ids = [
                item.id
                for item in queryset
                if q_norm in _normalize_text(item.ordem)
                or q_norm in _normalize_text(item.municipio)
                or q_norm in _normalize_text(item.projeto)
            ]
            queryset = queryset.filter(id__in=ids)
        if status:
            queryset = queryset.filter(status_ptm_atual__id=status)
        if municipio:
            municipio_norm = _normalize_text(municipio)
            ids = [item.id for item in queryset if municipio_norm in _normalize_text(item.municipio)]
            queryset = queryset.filter(id__in=ids)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["status_options"] = (
            StatusPTM.objects.order_by("nome").values_list("id", "nome").distinct()
        )
        context["can_create_ptm"] = self.request.user.is_superuser or bool(
            _allowed_municipios_for_user(self.request.user)
        )
        context["total_ptms_filtrados"] = context["paginator"].count
        return context


class DashboardView(TemplateView):
    template_name = "ptms/dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(_build_dashboard_context())
        return context


class NotaTecnicaMunicipioListView(TemplateView):
    template_name = "ptms/nota_tecnica_select.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        municipios = _municipios_com_ptms_para_usuario(self.request.user)
        context["municipios"] = municipios
        context["form"] = kwargs.get("form") or NotaTecnicaMunicipioForm(municipios=municipios)
        return context

    def post(self, request, *args, **kwargs):
        municipios = _municipios_com_ptms_para_usuario(request.user)
        form = NotaTecnicaMunicipioForm(request.POST, municipios=municipios)
        if form.is_valid():
            query = urlencode({"municipio": form.cleaned_data["municipio"]})
            return redirect(f"{reverse('nota_tecnica_detail')}?{query}")
        return self.render_to_response(self.get_context_data(form=form))


def nota_tecnica_detail(request):
    municipio = _resolve_municipio_para_nota(request.user, request.GET.get("municipio", ""))
    if not municipio:
        messages.error(request, "Selecione um municipio valido para gerar a nota tecnica.")
        return redirect("nota_tecnica_select")

    context = _build_nota_tecnica_context(municipio)
    if request.method == "POST":
        updated = 0
        with transaction.atomic():
            for section in context["sections"]:
                for item in section["items"]:
                    ptm = item["ptm"]
                    observacao = (request.POST.get(f"obs_{ptm.id}") or "").strip()
                    observacao_padrao = _nota_tecnica_default_observacao(ptm).strip()
                    nota = NotaTecnicaPTM.objects.filter(ptm=ptm).first()
                    if nota is None and observacao == observacao_padrao:
                        continue
                    if observacao == observacao_padrao:
                        nota.delete()
                        updated += 1
                        continue
                    if nota is None:
                        nota = NotaTecnicaPTM(ptm=ptm)
                    if nota.observacao != observacao or nota.updated_by_id != request.user.id:
                        nota.observacao = observacao
                        nota.updated_by = request.user
                        nota.save()
                        updated += 1
        messages.success(request, f"Nota tecnica atualizada com sucesso para {municipio}.")
        query = urlencode({"municipio": municipio})
        return redirect(f"{reverse('nota_tecnica_detail')}?{query}")

    return render(
        request,
        "ptms/nota_tecnica_detail.html",
        {
            **context,
            "updated_today": timezone.localdate(),
        },
    )


class PTMDetailView(DetailView):
    model = PTM
    template_name = "ptms/ptm_detail.html"
    context_object_name = "ptm"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ptm = self.object
        tab = self.request.GET.get("tab", "eventos")
        context["tab"] = tab
        context["eventos"] = ptm.eventos.select_related("status_ptm", "status_obra").all()
        context["termo_adesao"] = getattr(ptm, "termo_adesao_registro", None)
        context["termo_adesao_observacao_options"] = TermoAdesaoObservacao.objects.filter(ativo=True).order_by(
            "nome"
        )
        context["termo_adesao_responsavel_options"] = TermoAdesaoResponsavel.objects.filter(ativo=True).order_by(
            "nome"
        )
        context["pagamentos"] = ptm.pagamentos.all()
        context["status_ptm_options"] = StatusPTM.objects.filter(ativo=True).order_by("nome")
        context["status_obra_options"] = StatusObra.objects.filter(ativo=True).order_by("nome")
        context["vistorias"] = ptm.vistorias.all()
        context["prestacao"] = getattr(ptm, "prestacao_conta", None)
        context["prestacao_historico"] = (
            context["prestacao"].historico.all() if context["prestacao"] else []
        )
        context["observacoes"] = ptm.observacoes_enc.all()
        context["conclusoes"] = ptm.conclusoes_informais.all()
        context["documentos_publicos"] = ptm.documentos_publicos.all()
        context["status_analise_form"] = StatusAnaliseDocumentacaoForm(
            initial={"status_analise_documentacao": ptm.status_analise_documentacao}
        )
        context["can_edit_ptm"] = _user_can_edit_ptm(self.request.user, ptm)
        context["public_url"] = _public_ptm_url(self.request, ptm)
        context["resumo_financeiro"] = _build_resumo_financeiro(ptm)
        return context


class PTMPublicDetailView(DetailView):
    model = PTM
    template_name = "ptms/ptm_public_detail.html"
    context_object_name = "ptm"
    slug_field = "codigo_acesso_publico"
    slug_url_kwarg = "token"

    def get_object(self, queryset=None):
        token = self.kwargs.get("token", "").strip()
        if not token:
            raise Http404
        return get_object_or_404(
            PTM.objects.select_related(
                "status_ptm_atual",
                "status_obra_atual",
                "status_analise_documentacao",
            ),
            codigo_acesso_publico=token,
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ptm = self.object
        context["documento_publico_form"] = kwargs.get("documento_publico_form") or DocumentoPublicoPTMForm()
        context["documentos_publicos"] = ptm.documentos_publicos.all()
        return context


def ptm_public_upload(request, token):
    ptm = get_object_or_404(PTM, codigo_acesso_publico=token)
    if request.method != "POST":
        return redirect("ptm_public_detail", token=token)

    form = DocumentoPublicoPTMForm(request.POST, request.FILES)
    if form.is_valid():
        documento = form.save(commit=False)
        documento.ptm = ptm
        documento.save()
        messages.success(request, "Documento enviado com sucesso para analise.")
        return redirect("ptm_public_detail", token=token)

    return render(
        request,
        "ptms/ptm_public_detail.html",
        {
            "ptm": ptm,
            "documento_publico_form": form,
            "documentos_publicos": ptm.documentos_publicos.all(),
        },
    )


def ptm_generate_public_link(request, pk):
    ptm = get_object_or_404(PTM, pk=pk)
    blocked = _deny_edit_if_forbidden(request, ptm, tab=request.GET.get("tab", "eventos"))
    if blocked:
        return blocked
    if request.method == "POST":
        ptm.garantir_codigo_acesso_publico()
        ptm.save(update_fields=["codigo_acesso_publico"])
        messages.success(request, "Link publico gerado com sucesso.")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab={request.GET.get('tab', 'eventos')}")


def ptm_status_analise_update(request, pk):
    ptm = get_object_or_404(PTM, pk=pk)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="documentos_publicos")
    if blocked:
        return blocked
    if request.method == "POST":
        form = StatusAnaliseDocumentacaoForm(request.POST)
        if form.is_valid():
            ptm.status_analise_documentacao = form.cleaned_data["status_analise_documentacao"]
            ptm.save(update_fields=["status_analise_documentacao"])
            messages.success(request, "Status da analise da documentacao atualizado com sucesso.")
        else:
            messages.error(request, "Nao foi possivel atualizar o status da analise da documentacao.")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=documentos_publicos")


class PTMCreateView(CreateView):
    model = PTM
    form_class = PTMForm
    template_name = "ptms/ptm_form.html"

    def form_valid(self, form):
        allowed = _allowed_municipios_for_user(self.request.user)
        if allowed is not None:
            if not allowed:
                messages.error(
                    self.request,
                    "Seu usuario nao possui municipios vinculados para criar PTM.",
                )
                return self.form_invalid(form)
            municipio_form = _normalize_municipio(form.cleaned_data.get("municipio"))
            if municipio_form not in allowed:
                messages.error(
                    self.request,
                    "Voce so pode criar PTM para municipios vinculados ao seu usuario.",
                )
                return self.form_invalid(form)
        return super().form_valid(form)

    def get_success_url(self):
        messages.success(self.request, "PTM criado com sucesso.")
        return reverse("ptm_detail", kwargs={"pk": self.object.pk})


class PTMUpdateView(UpdateView):
    model = PTM
    form_class = PTMForm
    template_name = "ptms/ptm_form.html"

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        blocked = _deny_edit_if_forbidden(request, self.object)
        if blocked:
            return blocked
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        allowed = _allowed_municipios_for_user(self.request.user)
        if allowed is not None and _normalize_municipio(form.cleaned_data.get("municipio")) not in allowed:
            messages.error(
                self.request,
                "Voce so pode salvar PTM com municipio vinculado ao seu usuario.",
            )
            return self.form_invalid(form)
        return super().form_valid(form)

    def get_success_url(self):
        messages.success(self.request, "PTM atualizado com sucesso.")
        return reverse("ptm_detail", kwargs={"pk": self.object.pk})


class PTMDeleteView(DeleteView):
    model = PTM
    template_name = "ptms/ptm_confirm_delete.html"
    success_url = reverse_lazy("ptm_list")

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        blocked = _deny_edit_if_forbidden(request, self.object)
        if blocked:
            return blocked
        return super().dispatch(request, *args, **kwargs)


def evento_create(request, ptm_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="eventos")
    if blocked:
        return blocked
    if request.method == "POST":
        form = EventoPTMForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.ptm = ptm
            instance.save()
            messages.success(request, "Evento criado com sucesso.")
        else:
            errors = "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in form.errors.items())
            messages.error(request, f"Nao foi possivel criar evento. {errors}")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=eventos")


def evento_update(request, ptm_id, evento_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="eventos")
    if blocked:
        return blocked
    evento = get_object_or_404(EventoPTM, pk=evento_id, ptm=ptm)
    if request.method == "POST":
        form = EventoPTMForm(request.POST, instance=evento)
        if form.is_valid():
            form.save()
            messages.success(request, "Evento atualizado com sucesso.")
        else:
            errors = "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in form.errors.items())
            messages.error(request, f"Nao foi possivel atualizar evento. {errors}")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=eventos")


def evento_delete(request, ptm_id, evento_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="eventos")
    if blocked:
        return blocked
    evento = get_object_or_404(EventoPTM, pk=evento_id, ptm=ptm)
    if request.method == "POST":
        evento.delete()
        messages.success(request, "Evento excluido com sucesso.")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=eventos")


def pagamento_create(request, ptm_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="pagamentos")
    if blocked:
        return blocked
    if request.method == "POST":
        form = PagamentoPTMForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.ptm = ptm
            instance.save()
            messages.success(request, "Pagamento criado com sucesso.")
        else:
            errors = "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in form.errors.items())
            messages.error(request, f"Nao foi possivel criar pagamento. {errors}")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=pagamentos")


def pagamento_update(request, ptm_id, pagamento_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="pagamentos")
    if blocked:
        return blocked
    pagamento = get_object_or_404(PagamentoPTM, pk=pagamento_id, ptm=ptm)
    if request.method == "POST":
        form = PagamentoPTMForm(request.POST, instance=pagamento)
        if form.is_valid():
            form.save()
            messages.success(request, "Pagamento atualizado com sucesso.")
        else:
            errors = "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in form.errors.items())
            messages.error(request, f"Nao foi possivel atualizar pagamento. {errors}")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=pagamentos")


def pagamento_delete(request, ptm_id, pagamento_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="pagamentos")
    if blocked:
        return blocked
    pagamento = get_object_or_404(PagamentoPTM, pk=pagamento_id, ptm=ptm)
    if request.method == "POST":
        pagamento.delete()
        messages.success(request, "Pagamento excluido com sucesso.")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=pagamentos")


def vistoria_create(request, ptm_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="vistorias")
    if blocked:
        return blocked
    if request.method == "POST":
        form = VistoriaPTMForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            next_ordem = (
                ptm.vistorias.order_by("-ordem_vistoria").values_list("ordem_vistoria", flat=True).first() or 0
            ) + 1
            instance.ptm = ptm
            instance.ordem_vistoria = next_ordem
            instance.save()
            messages.success(request, "Vistoria criada com sucesso.")
        else:
            errors = "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in form.errors.items())
            messages.error(request, f"Nao foi possivel criar a vistoria. {errors}")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=vistorias")


def vistoria_update(request, ptm_id, vistoria_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="vistorias")
    if blocked:
        return blocked
    vistoria = get_object_or_404(VistoriaPTM, pk=vistoria_id, ptm=ptm)
    if request.method == "POST":
        post_data = request.POST.copy()
        raw_pct = (post_data.get("percentual_execucao") or "").strip()
        if raw_pct:
            try:
                pct = Decimal(raw_pct.replace(",", "."))
                if pct > Decimal("1"):
                    pct = pct / Decimal("100")
                post_data["percentual_execucao"] = str(pct)
            except (InvalidOperation, ValueError):
                pass
        form = VistoriaPTMForm(post_data, instance=vistoria)
        if form.is_valid():
            form.save()
            messages.success(request, "Vistoria atualizada com sucesso.")
        else:
            errors = "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in form.errors.items())
            messages.error(request, f"Nao foi possivel atualizar a vistoria. {errors}")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=vistorias")


def vistoria_delete(request, ptm_id, vistoria_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="vistorias")
    if blocked:
        return blocked
    vistoria = get_object_or_404(VistoriaPTM, pk=vistoria_id, ptm=ptm)
    if request.method == "POST":
        vistoria.delete()
        messages.success(request, "Vistoria excluida com sucesso.")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=vistorias")


def observacao_create(request, ptm_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="observacoes")
    if blocked:
        return blocked
    if request.method == "POST":
        form = ObservacaoEncaminhamentoForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.ptm = ptm
            instance.save()
            messages.success(request, "Observacao/encaminhamento criado com sucesso.")
        else:
            errors = "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in form.errors.items())
            messages.error(request, f"Nao foi possivel criar observacao/encaminhamento. {errors}")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=observacoes")


def observacao_update(request, ptm_id, observacao_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="observacoes")
    if blocked:
        return blocked
    observacao = get_object_or_404(ObservacaoEncaminhamentoPTM, pk=observacao_id, ptm=ptm)
    if request.method == "POST":
        form = ObservacaoEncaminhamentoForm(request.POST, instance=observacao)
        if form.is_valid():
            form.save()
            messages.success(request, "Observacao/encaminhamento atualizado com sucesso.")
        else:
            errors = "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in form.errors.items())
            messages.error(request, f"Nao foi possivel atualizar observacao/encaminhamento. {errors}")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=observacoes")


def observacao_delete(request, ptm_id, observacao_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="observacoes")
    if blocked:
        return blocked
    observacao = get_object_or_404(ObservacaoEncaminhamentoPTM, pk=observacao_id, ptm=ptm)
    if request.method == "POST":
        observacao.delete()
        messages.success(request, "Observacao/encaminhamento excluido com sucesso.")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=observacoes")


def conclusao_create(request, ptm_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="conclusoes")
    if blocked:
        return blocked
    if request.method == "POST":
        form = ConclusaoInformalForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.ptm = ptm
            instance.save()
            messages.success(request, "Conclusao informal criada com sucesso.")
        else:
            errors = "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in form.errors.items())
            messages.error(request, f"Nao foi possivel criar conclusao informal. {errors}")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=conclusoes")


def conclusao_update(request, ptm_id, conclusao_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="conclusoes")
    if blocked:
        return blocked
    conclusao = get_object_or_404(ConclusaoInformalPTM, pk=conclusao_id, ptm=ptm)
    if request.method == "POST":
        form = ConclusaoInformalForm(request.POST, instance=conclusao)
        if form.is_valid():
            form.save()
            messages.success(request, "Conclusao informal atualizada com sucesso.")
        else:
            errors = "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in form.errors.items())
            messages.error(request, f"Nao foi possivel atualizar conclusao informal. {errors}")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=conclusoes")


def conclusao_delete(request, ptm_id, conclusao_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="conclusoes")
    if blocked:
        return blocked
    conclusao = get_object_or_404(ConclusaoInformalPTM, pk=conclusao_id, ptm=ptm)
    if request.method == "POST":
        conclusao.delete()
        messages.success(request, "Conclusao informal excluida com sucesso.")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=conclusoes")


def prestacao_upsert(request, ptm_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="prestacao")
    if blocked:
        return blocked
    instance, _ = PrestacaoContaPTM.objects.get_or_create(ptm=ptm)
    if request.method == "POST":
        form = PrestacaoContaPTMForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Prestacao de contas atualizada com sucesso.")
        else:
            errors = "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in form.errors.items())
            messages.error(request, f"Nao foi possivel salvar prestacao. {errors}")
        return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=prestacao")
    else:
        form = PrestacaoContaPTMForm(instance=instance)

    return render(
        request,
        "ptms/child_form.html",
        {"form": form, "ptm": ptm, "title": "Editar Prestacao de Contas", "tab": "prestacao"},
    )


def termo_adesao_upsert(request, ptm_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="termo_adesao")
    if blocked:
        return blocked
    instance, _ = TermoAdesaoPTM.objects.get_or_create(ptm=ptm)
    if request.method == "POST":
        form = TermoAdesaoPTMForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Termo de adesao atualizado com sucesso.")
        else:
            errors = "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in form.errors.items())
            messages.error(request, f"Nao foi possivel salvar termo de adesao. {errors}")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=termo_adesao")


def termo_adesao_delete(request, ptm_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="termo_adesao")
    if blocked:
        return blocked
    termo = get_object_or_404(TermoAdesaoPTM, ptm=ptm)
    if request.method == "POST":
        termo.delete()
        messages.success(request, "Termo de adesao excluido com sucesso.")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=termo_adesao")


def prestacao_historico_create(request, ptm_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="prestacao")
    if blocked:
        return blocked
    if request.method == "POST":
        form = PrestacaoContaHistoricoForm(request.POST)
        if form.is_valid():
            prestacao, _ = PrestacaoContaPTM.objects.get_or_create(ptm=ptm)
            instance = form.save(commit=False)
            instance.prestacao = prestacao
            instance.save()
            messages.success(request, "Observacao da prestacao adicionada com sucesso.")
        else:
            errors = "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in form.errors.items())
            messages.error(request, f"Nao foi possivel adicionar observacao. {errors}")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=prestacao")


def prestacao_historico_update(request, ptm_id, historico_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="prestacao")
    if blocked:
        return blocked
    prestacao = get_object_or_404(PrestacaoContaPTM, ptm=ptm)
    historico = get_object_or_404(PrestacaoContaHistorico, pk=historico_id, prestacao=prestacao)
    if request.method == "POST":
        form = PrestacaoContaHistoricoForm(request.POST, instance=historico)
        if form.is_valid():
            form.save()
            messages.success(request, "Observacao da prestacao atualizada com sucesso.")
        else:
            errors = "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in form.errors.items())
            messages.error(request, f"Nao foi possivel atualizar observacao. {errors}")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=prestacao")


def prestacao_historico_delete(request, ptm_id, historico_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="prestacao")
    if blocked:
        return blocked
    prestacao = get_object_or_404(PrestacaoContaPTM, ptm=ptm)
    historico = get_object_or_404(PrestacaoContaHistorico, pk=historico_id, prestacao=prestacao)
    if request.method == "POST":
        historico.delete()
        messages.success(request, "Observacao da prestacao excluida com sucesso.")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=prestacao")


def prestacao_delete(request, ptm_id):
    ptm = get_object_or_404(PTM, pk=ptm_id)
    blocked = _deny_edit_if_forbidden(request, ptm, tab="prestacao")
    if blocked:
        return blocked
    prestacao = PrestacaoContaPTM.objects.filter(ptm=ptm).first()
    if request.method == "POST" and prestacao:
        prestacao.delete()
        messages.success(request, "Prestacao de contas excluida com sucesso.")
    return redirect(f"{reverse('ptm_detail', kwargs={'pk': ptm.pk})}?tab=prestacao")
