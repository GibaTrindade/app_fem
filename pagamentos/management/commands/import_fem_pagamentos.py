from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet

from pagamentos.models import PagamentoPTM
from ptms.models import PTM


def _norm(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).strip().lower()


def _sheet_by_name(workbook, expected: str) -> Worksheet:
    expected_norm = _norm(expected)
    for ws in workbook.worksheets:
        if _norm(ws.title) == expected_norm:
            return ws
    raise CommandError(f"Aba '{expected}' nao encontrada no arquivo.")


def _to_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted.date()
            if isinstance(converted, date):
                return converted
        except (OverflowError, ValueError, TypeError):
            return None
    return None


def _to_decimal(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0.00")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    try:
        clean = str(value).strip()
        if "," in clean and "." in clean and clean.rfind(",") > clean.rfind("."):
            clean = clean.replace(".", "").replace(",", ".")
        else:
            clean = clean.replace(",", ".")
        return Decimal(clean).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def _iter_pag_rows(ws: Worksheet):
    for row in range(2, ws.max_row + 1):
        ordem = _to_str(ws[f"A{row}"].value)
        if ordem:
            yield row, ordem


@dataclass
class Counters:
    processed: int = 0
    pagamentos: int = 0
    ptm_missing: int = 0


class Command(BaseCommand):
    help = "Importa apenas pagamentos da planilha FEM para PTMs ja existentes."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Caminho da planilha .xlsx/.xlsm")
        parser.add_argument("--limit", type=int, default=0, help="Limite de linhas para teste")

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"Arquivo nao encontrado: {file_path}")

        limit = max(0, int(options["limit"] or 0))
        wb = load_workbook(file_path, data_only=True, keep_vba=True)
        ws_pagamentos = _sheet_by_name(wb, "PAGAMENTOS")

        counters = Counters()
        for row, ordem in _iter_pag_rows(ws_pagamentos):
            if limit and counters.processed >= limit:
                break

            ptm = PTM.objects.filter(ordem=ordem).first()
            if ptm is None:
                counters.ptm_missing += 1
                counters.processed += 1
                continue

            with transaction.atomic():
                ptm.pagamentos.all().delete()
                counters.pagamentos += self._import_pagamentos(ws_pagamentos, row, ptm)

            counters.processed += 1
            if counters.processed % 25 == 0:
                self.stdout.write(f"Processados {counters.processed} PTMs...")

        self.stdout.write(self.style.SUCCESS("Importacao de pagamentos concluida."))
        self.stdout.write(
            f"Linhas processadas: {counters.processed} | pagamentos criados: {counters.pagamentos}"
        )
        if counters.ptm_missing:
            self.stdout.write(f"PTMs nao encontrados (ignorados): {counters.ptm_missing}")

    def _import_pagamentos(self, ws: Worksheet, row: int, ptm: PTM) -> int:
        created = 0
        normal_blocks = [
            ("1", "H", "I", "J", "K", "L", "M", "N", "O"),
            ("2", "P", "Q", "R", "S", "T", "U", "V", "W"),
            ("3", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE"),
            ("4", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM"),
        ]
        for parcela, c_sol, c_env, c_prev, c_real, c_pg, c_ob, c_emp, c_obs in normal_blocks:
            valor_real = ws[f"{c_real}{row}"].value
            if valor_real in (None, "") and ws[f"{c_pg}{row}"].value in (None, ""):
                continue
            PagamentoPTM.objects.create(
                ptm=ptm,
                parcela=parcela,
                tipo_registro="normal",
                dt_solicitacao=_to_date(ws[f"{c_sol}{row}"].value),
                dt_envio_pg=_to_date(ws[f"{c_env}{row}"].value),
                dt_pagamento=_to_date(ws[f"{c_pg}{row}"].value),
                valor_previsto=_to_decimal(ws[f"{c_prev}{row}"].value),
                valor_realizado=_to_decimal(valor_real),
                numero_ob=_to_str(ws[f"{c_ob}{row}"].value),
                numero_empenho=_to_str(ws[f"{c_emp}{row}"].value),
                observacao=_to_str(ws[f"{c_obs}{row}"].value),
            )
            created += 1

        extra_blocks = [
            ("AN", "AO", "AP", "AQ", "AR", "AS"),
            ("AT", "AU", "AV", "AW", "AX", "AY"),
        ]
        for c_parcela, c_real, c_pg, c_ob, c_emp, c_obs in extra_blocks:
            parcela_raw = _to_str(ws[f"{c_parcela}{row}"].value)
            if not parcela_raw and ws[f"{c_real}{row}"].value in (None, ""):
                continue
            parcela_num = parcela_raw.replace("ª", "").replace("a", "").strip() or "1"
            if parcela_num not in {"1", "2", "3", "4"}:
                parcela_num = "1"
            PagamentoPTM.objects.create(
                ptm=ptm,
                parcela=parcela_num,
                tipo_registro="extra",
                dt_pagamento=_to_date(ws[f"{c_pg}{row}"].value),
                valor_realizado=_to_decimal(ws[f"{c_real}{row}"].value),
                numero_ob=_to_str(ws[f"{c_ob}{row}"].value),
                numero_empenho=_to_str(ws[f"{c_emp}{row}"].value),
                observacao=_to_str(ws[f"{c_obs}{row}"].value),
            )
            created += 1

        return created
