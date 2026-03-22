from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet

from conclusao_informal.models import ConclusaoInformalPTM
from core.models import (
    AreaInvestimento,
    Municipio,
    Secretaria,
    StatusObra,
    StatusPTM,
    TermoAdesaoObservacao,
    TermoAdesaoResponsavel,
    TipoFEM,
)
from eventos.models import EventoPTM
from observacoes.models import ObservacaoEncaminhamentoPTM
from pagamentos.models import PagamentoPTM
from prestacao_contas.models import PrestacaoContaHistorico, PrestacaoContaPTM
from ptms.models import PTM, TermoAdesaoPTM
from vistorias.models import VistoriaPTM


def _norm(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).strip().lower()


def _sheet_by_name(workbook, expected: str) -> Worksheet:
    expected_norm = _norm(expected)
    for ws in workbook.worksheets:
        if _norm(ws.title) == expected_norm:
            return ws
    raise CommandError(f"Aba '{expected}' nao encontrada no arquivo.")


def _optional_sheet_by_name(workbook, expected: str) -> Worksheet | None:
    expected_norm = _norm(expected)
    for ws in workbook.worksheets:
        if _norm(ws.title) == expected_norm:
            return ws
    return None


def _to_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_str_limited(value, max_length: int) -> str:
    return _to_str(value)[:max_length]


def _to_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted.date()
            if isinstance(converted, date):
                return converted
        except (OverflowError, ValueError, TypeError):
            return None
    return None


def _to_decimal(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0.00")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    try:
        clean = str(value).strip()
        if "," in clean and "." in clean and clean.rfind(",") > clean.rfind("."):
            clean = clean.replace(".", "").replace(",", ".")
        else:
            clean = clean.replace(",", ".")
        return Decimal(clean).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def _to_percentage(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0.0000")
    try:
        return Decimal(str(value)).quantize(Decimal("0.0001"))
    except (InvalidOperation, ValueError):
        return Decimal("0.0000")


@dataclass(frozen=True)
class SheetLayout:
    header_row: int
    first_data_row: int
    first_col: int
    has_ordem: bool = True


def _column_index(letter: str) -> int:
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def _logical_column(layout: SheetLayout, letter: str) -> str:
    return get_column_letter(layout.first_col + _column_index(letter))


def _value(ws: Worksheet, layout: SheetLayout, row: int, letter: str):
    if not layout.has_ordem:
        if letter.upper() == "A":
            return None
        offset = _column_index(letter) - 1
        return ws.cell(row=row, column=layout.first_col + offset).value
    return ws[f"{_logical_column(layout, letter)}{row}"].value


def _detect_layout(ws: Worksheet) -> SheetLayout:
    for row in range(1, min(ws.max_row, 15) + 1):
        for col in range(1, min(ws.max_column, 12) + 1):
            value = ws.cell(row=row, column=col).value
            if _norm(_to_str(value)) != "ordem":
                continue
            first_data_row = row + 1
            while first_data_row <= ws.max_row and not _to_str(ws.cell(row=first_data_row, column=col).value):
                first_data_row += 1
            return SheetLayout(header_row=row, first_data_row=first_data_row, first_col=col, has_ordem=True)
    for row in range(1, min(ws.max_row, 10) + 1):
        header_1 = _norm(_to_str(ws.cell(row=row, column=1).value))
        header_2 = _norm(_to_str(ws.cell(row=row, column=2).value))
        header_3 = _norm(_to_str(ws.cell(row=row, column=3).value))
        if header_1 in {"regiao", "rd"} and header_2 == "municipio" and header_3 == "projeto":
            return SheetLayout(header_row=row, first_data_row=row + 1, first_col=1, has_ordem=False)
    raise CommandError(f"Nao foi possivel localizar o cabecalho 'ORDEM' na aba '{ws.title}'.")


def _iter_ptm_rows(ws: Worksheet, layout: SheetLayout, fallback_ordens: list[tuple[int, str]] | None = None):
    if not layout.has_ordem:
        if fallback_ordens is None:
            raise CommandError(f"Aba '{ws.title}' sem coluna ORDEM e sem referencia auxiliar.")
        for row, ordem in fallback_ordens:
            if row >= layout.first_data_row:
                yield row, ordem
        return
    for row in range(layout.first_data_row, ws.max_row + 1):
        ordem = _to_str(_value(ws, layout, row, "A"))
        if ordem:
            yield row, ordem


def _header_value(ws: Worksheet, layout: SheetLayout, letter: str) -> str:
    if not layout.has_ordem:
        if letter.upper() == "A":
            return ""
        offset = _column_index(letter) - 1
        value = ws.cell(row=layout.header_row, column=layout.first_col + offset).value
        return _to_str(value)
    value = ws[f"{_logical_column(layout, letter)}{layout.header_row}"].value
    return _to_str(value)


def _is_emenda_inf_layout(ws: Worksheet, layout: SheetLayout) -> bool:
    return _norm(_header_value(ws, layout, "F")) == "deputado"


def _is_emenda_termo_layout(ws: Worksheet, layout: SheetLayout) -> bool:
    return _norm(_header_value(ws, layout, "E")) == "cod"


def _prestacao_has_situacao(ws: Worksheet, layout: SheetLayout) -> bool:
    return _norm(_header_value(ws, layout, "H")) == "situacao da pc"


def _get_or_create_nome(model, value: str):
    value = _to_str(value)
    if not value:
        return None
    obj, _ = model.objects.get_or_create(nome=value)
    return obj


def _required_nome(model, value: str, default: str):
    obj = _get_or_create_nome(model, value)
    if obj is None:
        obj, _ = model.objects.get_or_create(nome=default)
    return obj


@dataclass
class Counters:
    ptms_created: int = 0
    ptms_updated: int = 0
    eventos: int = 0
    pagamentos: int = 0
    vistorias: int = 0
    prestacoes: int = 0
    prestacao_historico: int = 0
    observacoes: int = 0
    conclusoes: int = 0
    termos_adesao: int = 0


class Command(BaseCommand):
    help = "Importa a planilha FEM XLSM para o banco SQLite."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Caminho da planilha .xlsm")
        parser.add_argument(
            "--limit", type=int, default=0, help="Limita quantidade de PTMs importados (teste)"
        )
        parser.add_argument(
            "--skip-related",
            action="store_true",
            help="Importa apenas APOIO + INF GERAIS (sem abas filhas).",
        )
        parser.add_argument(
            "--only-events",
            action="store_true",
            help="Importa apenas EVENTOS para PTMs ja existentes.",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"Arquivo nao encontrado: {file_path}")

        limit = max(0, int(options["limit"] or 0))
        skip_related = options["skip_related"]
        only_events = options["only_events"]

        wb = load_workbook(file_path, data_only=True, keep_vba=True)
        ws_apoio = _optional_sheet_by_name(wb, "APOIO")
        ws_inf = _sheet_by_name(wb, "INF GERAIS")
        ws_eventos = _sheet_by_name(wb, "EVENTOS")
        layout_apoio = None
        if ws_apoio is not None:
            try:
                layout_apoio = _detect_layout(ws_apoio)
            except CommandError:
                layout_apoio = None
        layout_inf = _detect_layout(ws_inf)
        layout_eventos = _detect_layout(ws_eventos)
        row_pairs = list(
            _iter_ptm_rows(
                ws_inf,
                layout_inf,
                fallback_ordens=list(_iter_ptm_rows(ws_eventos, layout_eventos)) if not layout_inf.has_ordem else None,
            )
        )
        if not skip_related and not only_events:
            ws_termo_adesao = _sheet_by_name(wb, "TERMO DE ADESÃO")
            ws_pagamentos = _sheet_by_name(wb, "PAGAMENTOS")
            ws_vistorias = _sheet_by_name(wb, "VISTORIA")
            ws_prestacao = _sheet_by_name(wb, "PRESTACAO DE CONTAS")
            ws_obs = _sheet_by_name(wb, "OBS  ENC")
            ws_conclusao = _sheet_by_name(wb, "CONCLUSAO INFORMAL")
            layout_termo_adesao = _detect_layout(ws_termo_adesao)
            layout_pagamentos = _detect_layout(ws_pagamentos)
            layout_vistorias = _detect_layout(ws_vistorias)
            layout_prestacao = _detect_layout(ws_prestacao)
            layout_obs = _detect_layout(ws_obs)
            layout_conclusao = _detect_layout(ws_conclusao)

        counters = Counters()
        self._seed_catalogs(ws_apoio, layout_apoio, ws_inf, layout_inf, row_pairs)

        processed = 0
        for row, ordem in row_pairs:
            if limit and processed >= limit:
                break
            with transaction.atomic():
                if only_events:
                    ptm = PTM.objects.filter(ordem=ordem).first()
                    if ptm is None:
                        ptm, _ = self._upsert_ptm(ws_inf, layout_inf, row, ordem)
                    ptm.eventos.all().delete()
                    counters.eventos += self._import_eventos(ws_eventos, layout_eventos, row, ptm)
                else:
                    ptm, created = self._upsert_ptm(ws_inf, layout_inf, row, ordem)
                    if created:
                        counters.ptms_created += 1
                    else:
                        counters.ptms_updated += 1

                    if not skip_related:
                        TermoAdesaoPTM.objects.filter(ptm=ptm).delete()
                        ptm.eventos.all().delete()
                        ptm.pagamentos.all().delete()
                        ptm.vistorias.all().delete()
                        ptm.observacoes_enc.all().delete()
                        ptm.conclusoes_informais.all().delete()
                        PrestacaoContaPTM.objects.filter(ptm=ptm).delete()

                        counters.eventos += self._import_eventos(ws_eventos, layout_eventos, row, ptm)
                        counters.termos_adesao += self._import_termo_adesao(
                            ws_termo_adesao, layout_termo_adesao, row, ptm
                        )
                        counters.pagamentos += self._import_pagamentos(
                            ws_pagamentos, layout_pagamentos, row, ptm
                        )
                        counters.vistorias += self._import_vistorias(
                            ws_vistorias, layout_vistorias, row, ptm
                        )
                        p_count, ph_count = self._import_prestacao(
                            ws_prestacao, layout_prestacao, row, ptm
                        )
                        counters.prestacoes += p_count
                        counters.prestacao_historico += ph_count
                        counters.observacoes += self._import_observacoes(ws_obs, layout_obs, row, ptm)
                        counters.conclusoes += self._import_conclusoes(
                            ws_conclusao, layout_conclusao, row, ptm
                        )
            processed += 1
            if processed % 25 == 0:
                self.stdout.write(f"Processados {processed} PTMs...")

        self.stdout.write(self.style.SUCCESS("Importacao concluida."))
        self.stdout.write(
            f"PTMs criados: {counters.ptms_created} | atualizados: {counters.ptms_updated}"
        )
        if only_events:
            self.stdout.write(f"Eventos importados: {counters.eventos}")
        elif skip_related:
            self.stdout.write("Importacao de dados relacionados ignorada via --skip-related.")
        else:
            self.stdout.write(
                "Registros: "
                f"eventos={counters.eventos}, termos_adesao={counters.termos_adesao}, pagamentos={counters.pagamentos}, "
                f"vistorias={counters.vistorias}, prestacoes={counters.prestacoes}, "
                f"prestacao_historico={counters.prestacao_historico}, "
                f"obs_enc={counters.observacoes}, conclusoes={counters.conclusoes}"
            )

    def _seed_catalogs(
        self,
        ws_apoio: Worksheet | None,
        layout_apoio: SheetLayout | None,
        ws_inf: Worksheet,
        layout_inf: SheetLayout,
        row_pairs: list[tuple[int, str]],
    ):
        for value in ("NORMAL", "MULHER", "EMENDA"):
            TipoFEM.objects.get_or_create(nome=value)

        is_emenda_inf = _is_emenda_inf_layout(ws_inf, layout_inf)
        for row, _ in row_pairs:
            _get_or_create_nome(Municipio, _value(ws_inf, layout_inf, row, "C"))
            _get_or_create_nome(
                StatusPTM,
                _value(ws_inf, layout_inf, row, "N" if is_emenda_inf else "M"),
            )
            _get_or_create_nome(
                StatusObra,
                _value(ws_inf, layout_inf, row, "O" if is_emenda_inf else "N"),
            )
            _get_or_create_nome(
                Secretaria,
                _value(ws_inf, layout_inf, row, "S" if is_emenda_inf else "Q"),
            )
            _get_or_create_nome(
                AreaInvestimento,
                _value(ws_inf, layout_inf, row, "U" if is_emenda_inf else "S"),
            )

        if ws_apoio is None or layout_apoio is None:
            return

        for row in range(layout_apoio.first_data_row, ws_apoio.max_row + 1):
            _get_or_create_nome(StatusPTM, _value(ws_apoio, layout_apoio, row, "B"))
            _get_or_create_nome(StatusObra, _value(ws_apoio, layout_apoio, row, "C"))
            _get_or_create_nome(Municipio, _value(ws_apoio, layout_apoio, row, "I"))

    def _upsert_ptm(self, ws_inf: Worksheet, layout_inf: SheetLayout, row: int, ordem: str):
        if _is_emenda_inf_layout(ws_inf, layout_inf):
            tipo_fem = _required_nome(TipoFEM, "EMENDA", "EMENDA")
            status_ptm = _get_or_create_nome(StatusPTM, _value(ws_inf, layout_inf, row, "N"))
            status_obra = _get_or_create_nome(StatusObra, _value(ws_inf, layout_inf, row, "O"))
            area = _get_or_create_nome(AreaInvestimento, _value(ws_inf, layout_inf, row, "U"))
            secretaria = _get_or_create_nome(Secretaria, _value(ws_inf, layout_inf, row, "S"))
            defaults = {
                "regiao": _to_str(_value(ws_inf, layout_inf, row, "B")),
                "municipio": _to_str(_value(ws_inf, layout_inf, row, "C")),
                "projeto": _to_str(_value(ws_inf, layout_inf, row, "D")),
                "projeto_detalhado": _to_str(_value(ws_inf, layout_inf, row, "E")),
                "deputado": _to_str(_value(ws_inf, layout_inf, row, "F")),
                "numero_emenda": _to_str(_value(ws_inf, layout_inf, row, "H")),
                "tipo_fem": tipo_fem,
                "status_ptm_atual": status_ptm,
                "status_obra_atual": status_obra,
                "data_final": _to_date(_value(ws_inf, layout_inf, row, "G")),
                "data_aprovacao": _to_date(_value(ws_inf, layout_inf, row, "P")),
                "teto_fem": _to_decimal(_value(ws_inf, layout_inf, row, "K")),
                "investimento_total": _to_decimal(_value(ws_inf, layout_inf, row, "I")),
                "recurso_fem": _to_decimal(_value(ws_inf, layout_inf, row, "J")),
                "rendimentos_fem": _to_decimal(_value(ws_inf, layout_inf, row, "L")),
                "contrapartida": _to_decimal(_value(ws_inf, layout_inf, row, "M")),
                "ressalva": _to_str(_value(ws_inf, layout_inf, row, "R")),
                "secretaria": secretaria,
                "area_investimento": area,
                "conta_ptm": _to_str_limited(_value(ws_inf, layout_inf, row, "V"), 50),
                "descricao": "",
                "populacao_beneficiada": int(_value(ws_inf, layout_inf, row, "W"))
                if isinstance(_value(ws_inf, layout_inf, row, "W"), (int, float))
                else None,
            }
        else:
            tipo_fem = _required_nome(TipoFEM, _value(ws_inf, layout_inf, row, "F"), "NORMAL")
            status_ptm = _get_or_create_nome(StatusPTM, _value(ws_inf, layout_inf, row, "M"))
            status_obra = _get_or_create_nome(StatusObra, _value(ws_inf, layout_inf, row, "N"))
            area = _get_or_create_nome(AreaInvestimento, _value(ws_inf, layout_inf, row, "S"))
            secretaria = _get_or_create_nome(Secretaria, _value(ws_inf, layout_inf, row, "Q"))
            defaults = {
                "regiao": _to_str(_value(ws_inf, layout_inf, row, "B")),
                "municipio": _to_str(_value(ws_inf, layout_inf, row, "C")),
                "projeto": _to_str(_value(ws_inf, layout_inf, row, "D")),
                "projeto_detalhado": _to_str(_value(ws_inf, layout_inf, row, "E")),
                "deputado": "",
                "numero_emenda": "",
                "tipo_fem": tipo_fem,
                "status_ptm_atual": status_ptm,
                "status_obra_atual": status_obra,
                "data_final": _to_date(_value(ws_inf, layout_inf, row, "G")),
                "data_aprovacao": _to_date(_value(ws_inf, layout_inf, row, "O")),
                "teto_fem": _to_decimal(_value(ws_inf, layout_inf, row, "H")),
                "investimento_total": _to_decimal(_value(ws_inf, layout_inf, row, "I")),
                "recurso_fem": _to_decimal(_value(ws_inf, layout_inf, row, "J")),
                "rendimentos_fem": _to_decimal(_value(ws_inf, layout_inf, row, "K")),
                "contrapartida": _to_decimal(_value(ws_inf, layout_inf, row, "L")),
                "ressalva": _to_str(_value(ws_inf, layout_inf, row, "P")),
                "secretaria": secretaria,
                "area_investimento": area,
                "conta_ptm": _to_str_limited(_value(ws_inf, layout_inf, row, "T"), 50),
                "descricao": _to_str(_value(ws_inf, layout_inf, row, "AI")),
                "populacao_beneficiada": int(_value(ws_inf, layout_inf, row, "U"))
                if isinstance(_value(ws_inf, layout_inf, row, "U"), (int, float))
                else None,
            }

        return PTM.objects.update_or_create(ordem=ordem, defaults=defaults)

    def _import_eventos(self, ws: Worksheet, layout: SheetLayout, row: int, ptm: PTM) -> int:
        to_create: list[EventoPTM] = []
        latest_status_ptm = None
        latest_status_obra = None
        latest_data = date.min
        col = layout.first_col + _column_index("F")
        while col + 3 <= ws.max_column:
            descricao = _to_str(ws.cell(row=row, column=col).value)
            data_evento = _to_date(ws.cell(row=row, column=col + 1).value)
            status_ptm_nome = _to_str(ws.cell(row=row, column=col + 2).value)
            status_obra_nome = _to_str(ws.cell(row=row, column=col + 3).value)

            if descricao or data_evento or status_ptm_nome or status_obra_nome:
                status_ptm = _get_or_create_nome(StatusPTM, status_ptm_nome)
                status_obra = _get_or_create_nome(StatusObra, status_obra_nome)
                if data_evento and status_ptm and status_obra:
                    to_create.append(
                        EventoPTM(
                            ptm=ptm,
                            data_evento=data_evento,
                            descricao=descricao or "(sem descricao)",
                            status_ptm=status_ptm,
                            status_obra=status_obra,
                        )
                    )
                    if data_evento >= latest_data:
                        latest_data = data_evento
                        latest_status_ptm = status_ptm
                        latest_status_obra = status_obra
            col += 4
        if to_create:
            EventoPTM.objects.bulk_create(to_create, batch_size=200)
            ptm.status_ptm_atual = latest_status_ptm
            ptm.status_obra_atual = latest_status_obra
            ptm.save(update_fields=["status_ptm_atual", "status_obra_atual", "updated_at"])
        return len(to_create)

    def _import_termo_adesao(self, ws: Worksheet, layout: SheetLayout, row: int, ptm: PTM) -> int:
        if _is_emenda_termo_layout(ws, layout):
            sei = _to_str(_value(ws, layout, row, "E"))
            data = None
            responsavel = _get_or_create_nome(TermoAdesaoResponsavel, _value(ws, layout, row, "G"))
            observacao = _get_or_create_nome(TermoAdesaoObservacao, _value(ws, layout, row, "H"))
            secretaria = _to_str(_value(ws, layout, row, "I"))
        else:
            sei = _to_str(_value(ws, layout, row, "E"))
            data = _to_date(_value(ws, layout, row, "F"))
            responsavel = _get_or_create_nome(TermoAdesaoResponsavel, _value(ws, layout, row, "G"))
            observacao = _get_or_create_nome(TermoAdesaoObservacao, _value(ws, layout, row, "H"))
            secretaria = _to_str(_value(ws, layout, row, "I"))
        if not any([sei, data, responsavel, observacao]):
            return 0
        TermoAdesaoPTM.objects.create(
            ptm=ptm,
            sei=sei,
            data=data,
            responsavel=responsavel,
            observacao=observacao,
            secretaria=secretaria,
        )
        return 1

    def _import_pagamentos(self, ws: Worksheet, layout: SheetLayout, row: int, ptm: PTM) -> int:
        created = 0
        normal_blocks = [
            ("1", "H", "I", "J", "K", "L", "M", "N", "O"),
            ("2", "P", "Q", "R", "S", "T", "U", "V", "W"),
            ("3", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE"),
            ("4", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM"),
        ]
        for parcela, c_sol, c_env, c_prev, c_real, c_pg, c_ob, c_emp, c_obs in normal_blocks:
            valor_real = _value(ws, layout, row, c_real)
            if valor_real in (None, "") and _value(ws, layout, row, c_pg) in (None, ""):
                continue
            PagamentoPTM.objects.create(
                ptm=ptm,
                parcela=parcela,
                tipo_registro="normal",
                dt_solicitacao=_to_date(_value(ws, layout, row, c_sol)),
                dt_envio_pg=_to_date(_value(ws, layout, row, c_env)),
                dt_pagamento=_to_date(_value(ws, layout, row, c_pg)),
                valor_previsto=_to_decimal(_value(ws, layout, row, c_prev)),
                valor_realizado=_to_decimal(valor_real),
                numero_ob=_to_str_limited(_value(ws, layout, row, c_ob), 50),
                numero_empenho=_to_str_limited(_value(ws, layout, row, c_emp), 50),
                observacao=_to_str(_value(ws, layout, row, c_obs)),
            )
            created += 1

        extra_blocks = [
            ("AN", "AO", "AP", "AQ", "AR", "AS"),
            ("AT", "AU", "AV", "AW", "AX", "AY"),
        ]
        for c_parcela, c_real, c_pg, c_ob, c_emp, c_obs in extra_blocks:
            parcela_raw = _to_str(_value(ws, layout, row, c_parcela))
            if not parcela_raw and _value(ws, layout, row, c_real) in (None, ""):
                continue
            parcela_num = parcela_raw.replace("ª", "").replace("a", "").strip() or "1"
            if parcela_num not in {"1", "2", "3", "4"}:
                parcela_num = "1"
            PagamentoPTM.objects.create(
                ptm=ptm,
                parcela=parcela_num,
                tipo_registro="extra",
                dt_pagamento=_to_date(_value(ws, layout, row, c_pg)),
                valor_realizado=_to_decimal(_value(ws, layout, row, c_real)),
                numero_ob=_to_str_limited(_value(ws, layout, row, c_ob), 50),
                numero_empenho=_to_str_limited(_value(ws, layout, row, c_emp), 50),
                observacao=_to_str(_value(ws, layout, row, c_obs)),
            )
            created += 1
        return created

    def _import_vistorias(self, ws: Worksheet, layout: SheetLayout, row: int, ptm: PTM) -> int:
        created = 0
        col = layout.first_col + _column_index("F")
        idx = 1
        while col + 3 <= ws.max_column:
            c_sol = get_column_letter(col)
            c_resp = get_column_letter(col + 1)
            c_pct = get_column_letter(col + 2)
            c_obs = get_column_letter(col + 3)
            if ws[f"{c_sol}{row}"].value in (None, "") and ws[f"{c_resp}{row}"].value in (None, ""):
                col += 4
                idx += 1
                continue
            VistoriaPTM.objects.create(
                ptm=ptm,
                ordem_vistoria=idx,
                dt_solicitacao=_to_date(ws[f"{c_sol}{row}"].value),
                dt_resposta=_to_date(ws[f"{c_resp}{row}"].value),
                percentual_execucao=_to_percentage(ws[f"{c_pct}{row}"].value),
                observacao=_to_str(ws[f"{c_obs}{row}"].value),
            )
            created += 1
            col += 4
            idx += 1
        return created

    def _import_prestacao(self, ws: Worksheet, layout: SheetLayout, row: int, ptm: PTM) -> tuple[int, int]:
        has_situacao = _prestacao_has_situacao(ws, layout)
        base_has_data = any(
            _value(ws, layout, row, col) not in (None, "") for col in ("F", "G")
        )
        hist_pairs: list[tuple[str, str]] = []
        hist_start_letter = "I" if has_situacao else "H"
        col = layout.first_col + _column_index(hist_start_letter)
        while col + 1 <= ws.max_column:
            hist_pairs.append((get_column_letter(col), get_column_letter(col + 1)))
            col += 2
        hist_count = 0

        if not base_has_data and not any(
            ws[f"{obs_col}{row}"].value not in (None, "") for obs_col, _ in hist_pairs
        ):
            return 0, 0

        prestacao = PrestacaoContaPTM.objects.create(
            ptm=ptm,
            prazo_contas=_to_date(_value(ws, layout, row, "F")),
            data_prestacao=_to_date(_value(ws, layout, row, "G")),
            situacao=_to_str(_value(ws, layout, row, "H")) if has_situacao else "",
        )
        for obs_col, data_col in hist_pairs:
            obs = _to_str(ws[f"{obs_col}{row}"].value)
            data_registro = _to_date(ws[f"{data_col}{row}"].value)
            if obs or data_registro:
                PrestacaoContaHistorico.objects.create(
                    prestacao=prestacao,
                    data=data_registro,
                    observacao=obs or "(sem observacao)",
                )
                hist_count += 1
        return 1, hist_count

    def _import_observacoes(self, ws: Worksheet, layout: SheetLayout, row: int, ptm: PTM) -> int:
        created = 0
        col = layout.first_col + _column_index("F")
        while col + 1 <= ws.max_column:
            obs = _to_str(ws.cell(row=row, column=col).value)
            data_registro = _to_date(ws.cell(row=row, column=col + 1).value)
            if obs or data_registro:
                ObservacaoEncaminhamentoPTM.objects.create(
                    ptm=ptm,
                    data=data_registro,
                    observacao=obs or "(sem observacao)",
                )
                created += 1
            col += 2
        return created

    def _import_conclusoes(self, ws: Worksheet, layout: SheetLayout, row: int, ptm: PTM) -> int:
        created = 0
        blocks = [
            ("F", "G", "H", "I"),
            ("J", "K", "L", "M"),
            ("N", "O", "P", "Q"),
        ]
        for c_pct, c_data, c_contato, c_obs in blocks:
            pct = _value(ws, layout, row, c_pct)
            data_registro = _to_date(_value(ws, layout, row, c_data))
            contato = _to_str(_value(ws, layout, row, c_contato))
            obs = _to_str(_value(ws, layout, row, c_obs))
            if pct in (None, "") and not data_registro and not contato and not obs:
                continue
            ConclusaoInformalPTM.objects.create(
                ptm=ptm,
                percentual_declarado=_to_percentage(pct),
                data=data_registro,
                contato=contato,
                observacao=obs,
            )
            created += 1
        return created
